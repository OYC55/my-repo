import argparse
import datetime
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dart_client import (
    MAIN_ACCOUNTS,
    extract_main_accounts,
    find_company,
    get_financial_statement,
    load_corp_codes,
)

RATIO_LABELS = ["영업이익률(%)", "순이익률(%)", "ROE(%)", "ROA(%)", "부채비율(%)"]


def resolve_company(name, corp_codes):
    company, candidates = find_company(name, corp_codes)
    if company:
        return company

    if not candidates:
        print(f"'{name}'에 해당하는 회사를 찾을 수 없습니다.")
        sys.exit(1)

    print(f"\n'{name}'에 대해 여러 회사가 검색되었습니다. 번호를 선택하세요:")
    for i, c in enumerate(candidates, 1):
        stock = f" ({c['stock_code']})" if c["stock_code"] else ""
        print(f"  {i}. {c['corp_name']}{stock}")
    choice = input("선택 (숫자): ").strip()
    try:
        idx = int(choice) - 1
        return candidates[idx]
    except (ValueError, IndexError):
        print("잘못된 선택입니다.")
        sys.exit(1)


def compute_ratios(accounts):
    revenue = accounts.get("매출액")
    op_income = accounts.get("영업이익")
    net_income = accounts.get("당기순이익")
    assets = accounts.get("자산총계")
    liabilities = accounts.get("부채총계")
    equity = accounts.get("자본총계")

    def pct(numer, denom):
        if numer is None or denom in (None, 0):
            return None
        return round(numer / denom * 100, 2)

    return {
        "영업이익률(%)": pct(op_income, revenue),
        "순이익률(%)": pct(net_income, revenue),
        "ROE(%)": pct(net_income, equity),
        "ROA(%)": pct(net_income, assets),
        "부채비율(%)": pct(liabilities, equity),
    }


def fetch_data(companies, years):
    """{company_name: {year: {account: value, ...ratios}}}"""
    data = {}
    for company in companies:
        data[company["corp_name"]] = {}
        for year in years:
            print(f"조회 중: {company['corp_name']} {year}년...")
            statements = get_financial_statement(company["corp_code"], year)
            accounts = extract_main_accounts(statements)
            ratios = compute_ratios(accounts)
            data[company["corp_name"]][year] = {**accounts, **ratios}
    return data


def build_excel(data, companies, years, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "재무비교"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    section_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    company_names = [c["corp_name"] for c in companies]
    n_years = len(years)

    # Row 1: 회사명 (merged across each company's years)
    ws.cell(row=1, column=1, value="항목")
    col = 2
    for name in company_names:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n_years - 1)
        cell = ws.cell(row=1, column=col, value=name)
        cell.alignment = center
        col += n_years

    # Row 2: 연도
    col = 2
    for _ in company_names:
        for year in years:
            cell = ws.cell(row=2, column=col, value=f"{year}년")
            cell.alignment = center
            col += 1

    for row in (1, 2):
        for c in range(1, 2 + n_years * len(company_names)):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font

    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font

    row_idx = 3
    all_rows = list(MAIN_ACCOUNTS) + RATIO_LABELS
    for item in all_rows:
        if item == RATIO_LABELS[0]:
            pass  # 구분선 없이 이어서 표기
        cell = ws.cell(row=row_idx, column=1, value=item)
        cell.font = section_font

        col = 2
        for name in company_names:
            for year in years:
                value = data[name][year].get(item)
                out_cell = ws.cell(row=row_idx, column=col, value=value)
                if value is not None:
                    if item in RATIO_LABELS:
                        out_cell.number_format = "0.00"
                    else:
                        out_cell.number_format = "#,##0"
                col += 1
        row_idx += 1

    ws.column_dimensions["A"].width = 16
    for c in range(2, 2 + n_years * len(company_names)):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.freeze_panes = "B3"

    wb.save(output_path)
    print(f"\n엑셀 저장 완료: {output_path}")


def parse_args():
    parser = argparse.ArgumentParser(description="OpenDART 기반 재무제표 비교 분석기")
    parser.add_argument("companies", nargs="*", help="비교할 회사명 (최대 3개)")
    parser.add_argument("--years", nargs="+", type=int, help="조회 연도 (예: 2023 2024 2025)")
    parser.add_argument("--output", default="재무비교.xlsx", help="출력 엑셀 파일명")
    return parser.parse_args()


def main():
    args = parse_args()

    companies_input = args.companies
    while len(companies_input) < 3:
        name = input(f"비교할 회사명 {len(companies_input) + 1}/3: ").strip()
        if name:
            companies_input.append(name)
    companies_input = companies_input[:3]

    if args.years:
        years = sorted(args.years)
    else:
        current_year = datetime.date.today().year
        years = [current_year - 3, current_year - 2, current_year - 1]

    print("기업 코드 목록 로딩 중...")
    corp_codes = load_corp_codes()

    companies = [resolve_company(name, corp_codes) for name in companies_input]

    data = fetch_data(companies, years)
    build_excel(data, companies, years, args.output)


if __name__ == "__main__":
    main()
