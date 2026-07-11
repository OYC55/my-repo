from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analysis import RATIO_LABELS
from dart_client import MAIN_ACCOUNTS


def build_workbook(data, companies, years):
    """비교 데이터를 받아 openpyxl Workbook을 만들어 반환한다 (파일 저장은 호출부 책임)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "재무비교"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    section_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    company_names = [c["corp_name"] for c in companies]
    n_years = len(years)

    ws.cell(row=1, column=1, value="항목")
    col = 2
    for name in company_names:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n_years - 1)
        cell = ws.cell(row=1, column=col, value=name)
        cell.alignment = center
        col += n_years

    col = 2
    for _ in company_names:
        for year in years:
            cell = ws.cell(row=2, column=col, value=f"{year}년")
            cell.alignment = center
            col += 1

    for row in (1, 2):
        for c in range(1, 2 + n_years * len(company_names)):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font

    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font

    row_idx = 3
    all_rows = list(MAIN_ACCOUNTS) + RATIO_LABELS
    for item in all_rows:
        cell = ws.cell(row=row_idx, column=1, value=item)
        cell.font = section_font

        col = 2
        for name in company_names:
            for year in years:
                value = data[name][year].get(item)
                out_cell = ws.cell(row=row_idx, column=col, value=value)
                if value is not None:
                    out_cell.number_format = "0.00" if item in RATIO_LABELS else "#,##0"
                col += 1
        row_idx += 1

    ws.column_dimensions["A"].width = 16
    for c in range(2, 2 + n_years * len(company_names)):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.freeze_panes = "B3"

    return wb
